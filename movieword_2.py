import os
from tkinter import *
from tkinter import ttk
import tkinter as tk
import openpyxl
import re
from openpyxl import load_workbook

win = tk.Tk()
win.geometry("950x300")
win.title("熊的單字庫")

Ewordafter = tk.StringVar()
Cwordafter = tk.StringVar()
FromName = tk.StringVar()
Ewordbefore = tk.StringVar()
WorksheetName = tk.StringVar()
addEword = tk.StringVar()
addCword = tk.StringVar()
warnword = tk.StringVar()
SheetName = tk.StringVar()
filepath = tk.StringVar()
filename = os.path.abspath("wordbank.xlsx")

if not os.path.exists(filename):
    newcreate = openpyxl.Workbook()
    newcreate.save('wordbank.xlsx')
workbook = openpyxl.load_workbook('wordbank.xlsx')
count=len(workbook.sheetnames)

def searchEword():
    workbook = openpyxl.load_workbook('wordbank.xlsx')
    count=len(workbook.sheetnames)
    EwordafterString = ""
    CwordafterString = ""
    FromNameString = ""
    wordafternum = 0
    for k in range(0,count):
            sheet = workbook.worksheets[k]
            for i in range(1,sheet.max_row+1):
                for j in range(1,sheet.max_column+1):
                    if Ewordbefore.get() == str(sheet.cell(row=i,column=j).value):
                        wordafternum = wordafternum + 1
                        if wordafternum == 1:
                            EwordafterString = str(sheet.cell(row=i,column=j).value)
                            CwordafterString = str(sheet.cell(row=i,column=j+1).value)
                            FromNameString = str(workbook.sheetnames[k])
                        else:
                            EwordafterString = EwordafterString + "\n" + str(sheet.cell(row=i,column=j).value)
                            CwordafterString = CwordafterString + "\n" + str(sheet.cell(row=i,column=j+1).value)
                            FromNameString = FromNameString + "\n" + str(workbook.sheetnames[k])
                        break
    
    Ewordafter.set(EwordafterString)
    Cwordafter.set(CwordafterString)
    FromName.set(FromNameString)
    if EwordafterString == "":
        Ewordafter.set("查無此單字")

def addword():
    WorksheetNameString = WorksheetName.get()
    addEwordString = addEword.get()
    addCwordString = addCword.get()
    sheetneedCreate = 1
    entryfinish = 0
    entryerror = 0
    wb = load_workbook('wordbank.xlsx')
    count=len(wb.sheetnames)

    if (WorksheetNameString == "") or (addEwordString == "") or (addCwordString == ""):
        entryfinish = 1

    WorksheetNameCheck = re.compile(r'^\s|\s$')  #開頭與結尾不能有空格
    WorksheetNameJudge = WorksheetNameCheck.search(WorksheetNameString)
    addEwordCheck = re.compile(r'[^a-z]|[^A-Z] | ^\s|\s$')  #開頭與結尾不能有空格或不能有非英文
    addEwordJudge = addEwordCheck.search(addEwordString)
    addCwordCheck = re.compile(r'^\s|\s$')   #開頭與結尾不能有空格
    addCwordJudge = addCwordCheck.search(addCwordString)

    if (WorksheetNameJudge != None) or (addCwordJudge != None) or (addEwordJudge != None):
        entryerror = 1
        label8.config(fg="red")
        warnword.set("輸入格式錯誤!")

    if (entryerror == 1) or (entryfinish == 1):
        if (entryerror == 1) and (entryfinish == 0):
            label8.config(fg="red")
            warnword.set("輸入格式錯誤!")
        if (entryerror == 0) and (entryfinish == 1):
            label8.config(fg="red")
            warnword.set("輸入不完整!")
        if (entryerror == 1) and (entryfinish == 1):
            label8.config(fg="red")
            warnword.set("輸入不完整與輸入格式錯誤!")

    if (entryfinish == 0) and (entryerror == 0):
        for i in range(0,count):
            if WorksheetNameString == wb.sheetnames[i]:
                sheetneedCreate = 0
                sheet = wb.worksheets[i]
                dataWrite = [addEwordString,addCwordString]
                sheet.append(dataWrite)
                label8.config(fg="blue")
                warnword.set("工作表已存在，單字儲存成功!")

        if sheetneedCreate == 1:
            ws1 = wb.create_sheet(WorksheetNameString, count)
            sheet = wb.worksheets[count]
            dataWrite = [addEwordString,addCwordString]
            sheet.append(dataWrite)
            label8.config(fg="blue")
            warnword.set("已創建新的工作表，單字儲存成功!")
        
        count=len(wb.sheetnames)
        wb.save('wordbank.xlsx')
        wb.close()

def renewsheetname():
    SheetNameString =""
    workbook = openpyxl.load_workbook('wordbank.xlsx')
    count=len(workbook.sheetnames)
    for i in range(0, count):
        if i == 0:
            SheetNameString = workbook.sheetnames[i]
        else:    
            SheetNameString = SheetNameString + "," + workbook.sheetnames[i]
        if (i!=0) and (i % 10) == 0:
            SheetNameString = SheetNameString + "\n"   #每 10 個換行
    SheetName.set(SheetNameString)


tabsystem = ttk.Notebook(win)
tab1 = Frame(tabsystem)
tab2 = Frame(tabsystem)
tab3 = Frame(tabsystem)

tabsystem.add(tab1, text='新增單字')
tabsystem.add(tab2, text='查詢單字')
tabsystem.add(tab3, text='工作表名稱庫')
tabsystem.pack(expand=1, fill="both")

#-------------------------分頁 1---------------------------

labelNone4 = tk.Label(tab1, text="")
label5 = tk.Label(tab1, text="工作表名稱：", font=("微軟正黑體", 18))
entry2 = tk.Entry(tab1, textvariable=WorksheetName, font=("微軟正黑體", 18))
label6 = tk.Label(tab1, text="增加的英文單字：", font=("微軟正黑體", 18))
entry3 = tk.Entry(tab1, textvariable=addEword, font=("微軟正黑體", 18))
label7 = tk.Label(tab1, text="中文意思：", font=("微軟正黑體", 18))
entry4 = tk.Entry(tab1, textvariable=addCword, font=("微軟正黑體", 18))
labelNone3 = tk.Label(tab1, text="")
button2 = tk.Button(tab1, text="增加至單字庫", font=("微軟正黑體", 18, "bold"), bg="CadetBlue1", fg="black", command=addword)
label8 = tk.Label(tab1, textvariable=warnword, font=("微軟正黑體", 14))
label10 = tk.Label(tab1, textvariable=filepath, font=("微軟正黑體", 12))
filepath.set("目前讀取檔案路徑：" + filename)

labelNone4.grid(row=0, column=0)
label5.grid(row=1, column=0, sticky="e")
entry2.grid(row=1, column=1)
label6.grid(row=2, column=0, sticky="e")
entry3.grid(row=2, column=1)
label7.grid(row=3, column=0, sticky="e")
entry4.grid(row=3, column=1)
labelNone3.grid(row=4, column=0)
button2.grid(row=5, column=1, sticky="w")
label8.place(x=190, y=210)
label10.place(x=50, y=250)

#-------------------------分頁 2---------------------------

frame1 = tk.Frame(tab2)
frame1.pack()
label1 = tk.Label(frame1, text="想查詢的單字：", font=("微軟正黑體", 18))
entry = tk.Entry(frame1, textvariable=Ewordbefore, font=("微軟正黑體", 18))
labelNone1 = tk.Label(frame1, text=" ", font=("微軟正黑體", 18))
button1 = tk.Button(frame1, text="查詢", font=("微軟正黑體", 18, "bold"), bg="yellow", fg="blue", command=searchEword)
label1.grid(row=0, column=0)
entry.grid(row=0, column=1)
labelNone1.grid(row=0, column=2)
button1.grid(row=0, column=3)

labelNone2 = tk.Label(tab2, text="")
labelNone2.pack()

frame2 = tk.Frame(tab2)
frame2.pack()
labeltitle1 = tk.Label(frame2, text="英文單字".ljust(18), font=("微軟正黑體", 18), bg="black", fg="white")
labeltitle2 = tk.Label(frame2, text="中文".ljust(80), font=("微軟正黑體", 18), bg="black", fg="white")
labeltitle3 = tk.Label(frame2, text="出處".ljust(30), font=("微軟正黑體", 18), bg="black", fg="white")
label2 = tk.Label(frame2, textvariable=Ewordafter, font=("微軟正黑體", 14))
label3 = tk.Label(frame2, textvariable=Cwordafter, font=("微軟正黑體",14))
label4 = tk.Label(frame2, textvariable=FromName, font=("微軟正黑體", 14))
labeltitle1.grid(row=0, column=0)
labeltitle2.grid(row=0, column=1)
labeltitle3.grid(row=0, column=2)
label2.grid(row=1, column=0, sticky="w")
label3.grid(row=1, column=1, sticky="w")
label4.grid(row=1, column=2, sticky="w")

#-------------------------分頁 3---------------------------

button3 = tk.Button(tab3, text="更新", font=("微軟正黑體", 18), command=renewsheetname)
label9 = tk.Label(tab3, textvariable=SheetName, font=("微軟正黑體", 12))
button3.pack()
label9.pack()

win.mainloop()
