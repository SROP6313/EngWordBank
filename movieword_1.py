# -*- coding: utf-8 -*-
"""
Created on Wed Jan 26 21:44:10 2022

@author: user
"""

import openpyxl
workbook = openpyxl.load_workbook('moviewo.xlsx')
count=len(workbook.sheetnames)
while(1):
    word = input("輸入要查的單字 (關閉程式請打 1)：")
    if word == str(1):
        break
    else:
        for k in range(0,count):
            sheet = workbook.worksheets[k]
            for i in range(1,sheet.max_row+1):
                for j in range(1,sheet.max_column+1):
                    if word == str(sheet.cell(row=i,column=j).value):
                        print(sheet.cell(row=i,column=j).value + " ==> " + sheet.cell(row=i,column=j+1).value + "   from " + workbook.sheetnames[k]) 
                        break